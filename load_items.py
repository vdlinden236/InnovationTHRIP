import datetime
import pymssql
import openpyxl as opx


#Create the dictionary
def contract_data(file_name, orgid, userid):
    now = datetime.date.today()

    #excel_book = opx.load_workbook('{}'.format(WorkBook))
    excel_book = opx.load_workbook(file_name)
    sheet_1 = excel_book.get_sheet_by_name('Sheet1')

    N_Columns = sheet_1.max_column
    N_Rows = sheet_1.max_row
    Extracted_Data = {}
    Extracted_Data['Items'] = []
    t_qty = 0
    qty = 0
    pos = []
    n_pos = []

    #Calculate qty of items
    if sheet_1.cell(row=1, column=1).value == 'Ref':
        for i in range(2, N_Rows+1):
            if sheet_1.cell(row=i, column=3).value is not None:
                t_qty += int(sheet_1.cell(row=i, column=3).value)
                if int(sheet_1.cell(row=i, column=3).value) > 1:
                    qty += int(sheet_1.cell(row=i, column=3).value)
                    pos.append(i)
                else:
                    n_pos.append(i)

    #Append a list for each item
    for i in range(t_qty):
        Extracted_Data['Items'].append([])

    ans = 0
    for i in pos:
        cnt = 1
        for a in range(int(sheet_1.cell(row=i, column=3).value)):
            Extracted_Data['Items'][ans + a].append(str(sheet_1.cell(row=i, column=1).value) + "_" + str(cnt).replace(u'\xa0', u' '))
            Extracted_Data['Items'][ans + a].append(str(sheet_1.cell(row=i, column=2).value).replace(u'\xa0', u' '))
            cnt += 1
            for cols in range(3, N_Columns+1):
                try:
                    Extracted_Data['Items'][ans + a].append(str(round(sheet_1.cell(row=i, column=cols).value, 2)).replace(u'\xa0', u'').replace('R', ''))
                except TypeError:
                    Extracted_Data['Items'][ans + a].append(str(sheet_1.cell(row=i, column=cols).value).replace(u'\xa0', u'').replace('R', ''))
        ans += (1 + a)


    for i in n_pos:
        Extracted_Data['Items'][ans].append(str(sheet_1.cell(row=i, column=1).value).replace(u'\xa0', u' '))
        Extracted_Data['Items'][ans].append(str(sheet_1.cell(row=i, column=2).value).replace(u'\xa0', u' '))
        for cols in range(3, N_Columns+1):
            try:
                Extracted_Data['Items'][ans].append(str(round(sheet_1.cell(row=i, column=cols).value, 2)).replace(u'\xa0', u'').replace('R', ''))
            except TypeError:
                Extracted_Data['Items'][ans].append(str(sheet_1.cell(row=i, column=cols).value).replace(u'\xa0', u'').replace('R', ''))
        ans += 1


    #Connect to database
    host = "197.189.232.50"
    username = "FE-User"
    password = "Fourier.01"
    database = "PGAluminium"

    conn = pymssql.connect(host, username, password, database)
    cursor = conn.cursor()


    ActStat = cursor.execute("SELECT a.StatusId from ActivityStatuses a left join Activities s on a.ActivityId=s.ActivityId where s.Level = 2 and a.IsDefault=1 and a.OrgId = {}".format(orgid))

    ActStat = cursor.fetchone()
    ActStat = ActStat[0]


    for i in range(len(Extracted_Data['Items'])):
        cursor.execute("INSERT INTO ContractItems (ContractId, OrgId, LineReferenceNumber, Description, Value, TimeStamped, UserId, ActivityStatusId, DurationEstMins) VALUES ({}, {}, '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(contractid, orgid, Extracted_Data["Items"][i][0],Extracted_Data["Items"][i][1],Extracted_Data["Items"][i][3],now,userid,ActStat,Extracted_Data["Items"][i][4]))
    
    conn.commit()

    return Extracted_Data



#print(contract_data())

"""
#Connect to database
host = "41.203.23.36"
username = "FE-User"
password = "Fourier.01"
database = "PGAluminium"

conn = pymssql.connect(host, username, password, database)
cursor = conn.cursor()

cursor.execute("Truncate table Contracts")
cursor.execute("Truncate table ContractItems")
conn.commit()
"""