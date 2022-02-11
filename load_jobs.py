import datetime
import pymssql
import openpyxl as opx
from openpyxl.styles import NamedStyle

#Create the dictionary
def contract_data(file_name, orgid, userid):
    now = datetime.date.today()
    date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')
    #excel_book = opx.load_workbook('{}'.format(WorkBook))
    excel_book = opx.load_workbook(file_name)
    #excel_book = opx.load_workbook('static/reports/SFA.xlsx')
    print(excel_book)
    sheet_1 = excel_book['Sheet1']

    N_Columns = sheet_1.max_column
    N_Rows = sheet_1.max_row
    Extracted_Data = {}
    Extracted_Data['Items'] = []
    rows = 1

    for cell in sheet_1['A']:
        cell.style = date_style

    for cell in sheet_1['B']:
        cell.style = date_style

    for i in range(0, N_Rows-1):

        if sheet_1.cell(row=i+1, column=1).value is None:
            break
        else:
            Extracted_Data['Items'].append([])

    for i in range(0, N_Rows-1):
        Extracted_Data['Items'][i].append(sheet_1.cell(row=i + 2, column=1).value.replace(u'\xa0', u' '))
        Extracted_Data['Items'][i].append(sheet_1.cell(row=i+2, column=2).value.replace(u'\xa0', u' '))
        Extracted_Data['Items'][i].append(str(sheet_1.cell(row=i+2, column=3).value).replace(u'\xa0', u' '))
        Extracted_Data['Items'][i].append(str(sheet_1.cell(row=i+2, column=4).value).replace(u'\xa0', u' '))
        Extracted_Data['Items'][i].append(str(sheet_1.cell(row=i+2, column=5).value).replace(u'\xa0', u' '))
        Extracted_Data['Items'][i].append(str(sheet_1.cell(row=i + 2, column=6).value).replace(u'\xa0', u' '))

    def remove_empty_keys(d):
        for k in d.keys():
            if d[k] is None:
                del d[k]

        remove_empty_keys(Extracted_Data)


    #Connect to database
    host = "197.189.232.50"
    username = "FE-User"
    password = "Fourier.01"
    database = "PGAluminium"

    conn = pymssql.connect(host, username, password, database)
    cursor = conn.cursor()

    Status = "Start"
    Finish = ""
    Notes = ""

    for i in range(0, len(Extracted_Data['Items'])):
        cursor.execute("""
                        INSERT INTO Contracts 
                        (OrgId, ContractReference, Description, DueDate, AssignedUserId, ContactPerson, ContactNumber1, 
                        ContactNumber2, Notes, Value, OrderedDate, Timestamp) 
                        VALUES ({}, '{}', '{}', CONVERT(date, '{}', 103), {}, '{}', '{}', '{}', '{}', {}, CONVERT(date, '{}', 103), '{}')
                       """
                       .format(orgid, Extracted_Data['Items'][i][2], Extracted_Data['Items'][i][3], Extracted_Data['Items'][i][1], userid, Extracted_Data['Items'][i][4], Status, Finish, Notes, Extracted_Data['Items'][i][5], Extracted_Data['Items'][i][0], now))
    
    conn.commit()
    return Extracted_Data



#print(contract_data(14, 86))

"""
#Connect to database
host = "41.203.23.36"
username = "FE-User"
password = "Fourier.01"
database = "PGAluminium"

conn = pymssql.connect(host, username, password, database)
cursor = conn.cursor()

cursor.execute("Truncate table Contracts")
cursor.execute("Truncate table ContractItems")
conn.commit()
"""